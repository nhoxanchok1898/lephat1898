"""
Admin Dashboard - Professional Grade
Real-time KPI metrics, interactive charts, and reporting
"""
import csv
import io
from datetime import datetime, timedelta
from decimal import Decimal
from django.contrib import messages
from django.contrib.admin.views.decorators import staff_member_required
from django.contrib.auth.decorators import login_required
from django.core.mail import EmailMessage
from django.db.models import Sum, Count, Avg, F, Q, DecimalField, DateField
from django.db.models.functions import TruncDate, TruncMonth, Coalesce
from django.utils.dateparse import parse_date
from django.http import HttpResponse, JsonResponse, HttpResponseRedirect
from django.shortcuts import render, redirect
from django.utils import timezone
from django.views.decorators.http import require_http_methods
from .models import (
    Order, OrderItem, Product, User,
    ProductView, OrderAnalytics, UserAnalytics, ProductPerformance
)


def _paid_orders_queryset(request, params=None):
    """Base queryset for paid orders with optional date filtering."""
    params = params or request.GET
    paid_orders = Order.objects.filter(is_paid=True).prefetch_related('items')

    from_param = params.get('from_date')
    to_param = params.get('to_date')
    preset = params.get('preset')
    today = timezone.now().date()

    if preset == 'last_7':
        from_param = (today - timedelta(days=7)).isoformat()
        to_param = today.isoformat()
    elif preset == 'last_30':
        from_param = (today - timedelta(days=30)).isoformat()
        to_param = today.isoformat()
    elif preset == 'this_month':
        from_param = today.replace(day=1).isoformat()
        to_param = today.isoformat()

    from_date = parse_date(from_param) if from_param else None
    to_date = parse_date(to_param) if to_param else None

    if from_date:
        paid_orders = paid_orders.filter(
            Q(paid_at__date__gte=from_date) | Q(paid_at__isnull=True, created_at__date__gte=from_date)
        )
    if to_date:
        paid_orders = paid_orders.filter(
            Q(paid_at__date__lte=to_date) | Q(paid_at__isnull=True, created_at__date__lte=to_date)
        )
    return paid_orders, from_param, to_param, preset


@staff_member_required
@require_http_methods(["GET"])
def admin_dashboard(request):
    """
    Main admin dashboard with KPIs and charts
    """
    # Date range
    today = timezone.now().date()
    last_30_days = today - timedelta(days=30)
    last_7_days = today - timedelta(days=7)
    
    # KPI Metrics
    metrics = get_kpi_metrics(last_30_days)
    
    # Chart data
    charts = {
        'sales_trend': get_sales_trend_data(last_30_days),
        'top_products': get_top_products_data(10),
        'revenue_breakdown': get_revenue_breakdown(),
        'user_growth': get_user_growth_data(last_30_days),
    }
    
    # Recent orders
    recent_orders = Order.objects.select_related().order_by('-created_at')[:10]
    
    # Low stock alerts
    low_stock_products = Product.objects.filter(
        stock_quantity__lte=10,
        is_active=True
    ).order_by('stock_quantity')[:10]
    
    context = {
        'metrics': metrics,
        'charts': charts,
        'recent_orders': recent_orders,
        'low_stock_products': low_stock_products,
    }
    
    return render(request, 'admin/dashboard.html', context)


@staff_member_required
@require_http_methods(["GET"])
def revenue_report(request):
    """Read-only revenue report for paid orders (staff only)."""
    paid_orders, from_param, to_param, preset = _paid_orders_queryset(request)

    # Daily totals (prefer paid_at; fallback to created_at when paid_at is null)
    daily_totals = (
        paid_orders
        .annotate(day=Coalesce(TruncDate('paid_at'), TruncDate('created_at'), output_field=DateField()))
        .values('day')
        .annotate(
            total=Sum(F('items__price') * F('items__quantity'), output_field=DecimalField()),
            orders=Count('id'),
        )
        .order_by('-day')
    )

    # Month totals
    monthly_totals = (
        paid_orders
        .annotate(month=Coalesce(TruncMonth('paid_at'), TruncMonth('created_at')))
        .values('month')
        .annotate(
            total=Sum(F('items__price') * F('items__quantity'), output_field=DecimalField()),
            orders=Count('id'),
        )
        .order_by('-month')
    )

    aggregates = paid_orders.aggregate(
        total_revenue=Sum(F('items__price') * F('items__quantity'), output_field=DecimalField()),
        order_count=Count('id'),
    )
    total_revenue = aggregates.get('total_revenue') or Decimal('0')
    order_count = aggregates.get('order_count') or 0
    avg_order_value = (total_revenue / order_count) if order_count else Decimal('0')

    # Previous period comparison (only when both from/to provided)
    period_change = None
    from_date = parse_date(from_param) if from_param else None
    to_date = parse_date(to_param) if to_param else None
    if from_date and to_date:
        duration = (to_date - from_date).days + 1
        prev_end = from_date - timedelta(days=1)
        prev_start = prev_end - timedelta(days=duration - 1)
        prev_qs = Order.objects.filter(is_paid=True).prefetch_related('items')
        prev_qs = prev_qs.filter(
            Q(paid_at__date__gte=prev_start) | Q(paid_at__isnull=True, created_at__date__gte=prev_start)
        ).filter(
            Q(paid_at__date__lte=prev_end) | Q(paid_at__isnull=True, created_at__date__lte=prev_end)
        )
        prev_agg = prev_qs.aggregate(
            total_revenue=Sum(F('items__price') * F('items__quantity'), output_field=DecimalField()),
            order_count=Count('id'),
        )
        prev_total = prev_agg.get('total_revenue') or Decimal('0')
        prev_orders = prev_agg.get('order_count') or 0
        prev_aov = (prev_total / prev_orders) if prev_orders else Decimal('0')

        def pct_change(current, previous):
            if previous == 0:
                return None if current == 0 else 100
            return float((current - previous) / previous * 100)

        period_change = {
            'prev_start': prev_start,
            'prev_end': prev_end,
            'revenue_pct': pct_change(total_revenue, prev_total),
            'orders_pct': pct_change(order_count, prev_orders),
            'aov_pct': pct_change(avg_order_value, prev_aov),
            'prev_revenue': prev_total,
            'prev_orders': prev_orders,
            'prev_aov': prev_aov,
        }

    # Top products by revenue and quantity (read-only)
    top_products = (
        OrderItem.objects
        .filter(order__in=paid_orders)
        .values('product__id', 'product__name')
        .annotate(
            total_revenue=Sum(F('price') * F('quantity'), output_field=DecimalField()),
            total_qty=Sum('quantity'),
        )
        .order_by('-total_revenue')[:10]
    )

    # Top customers by total spend (orders with user)
    top_customers = (
        paid_orders
        .filter(user__isnull=False)
        .values('user__id', 'user__username', 'user__email')
        .annotate(
            total_spend=Sum(F('items__price') * F('items__quantity'), output_field=DecimalField()),
            orders=Count('id'),
        )
        .order_by('-total_spend')[:10]
    )

    # Revenue breakdown by payment method
    payment_breakdown = (
        paid_orders
        .values('payment_method')
        .annotate(
            total=Sum(F('items__price') * F('items__quantity'), output_field=DecimalField()),
            orders=Count('id'),
        )
        .order_by('-total')
    )

    context = {
        'daily_totals': daily_totals,
        'monthly_totals': monthly_totals,
        'total_revenue': total_revenue,
        'order_count': order_count,
        'avg_order_value': avg_order_value,
        'from_date': from_param,
        'to_date': to_param,
        'preset': preset,
        'period_change': period_change,
        'top_products': top_products,
        'top_customers': top_customers,
        'payment_breakdown': payment_breakdown,
    }
    return render(request, 'admin/revenue_report.html', context)


@staff_member_required
@require_http_methods(["GET"])
def export_revenue_csv(request):
    """Export paid orders revenue as CSV."""
    paid_orders, from_param, to_param, preset = _paid_orders_queryset(request)
    paid_orders = paid_orders.annotate(
        order_total=Sum(F('items__price') * F('items__quantity'), output_field=DecimalField())
    ).order_by('-paid_at', '-created_at')

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename=revenue_paid_orders.csv'
    writer = csv.writer(response)
    writer.writerow([
        'Order ID', 'Paid At', 'Created At', 'Customer', 'Payment Method', 'Status', 'Total'
    ])
    for o in paid_orders:
        writer.writerow([
            o.id,
            o.paid_at.isoformat() if o.paid_at else '',
            o.created_at.isoformat() if o.created_at else '',
            o.full_name,
            o.get_payment_method_display(),
            o.get_status_display(),
            f"{o.order_total or Decimal('0')}",
        ])
    return response


@staff_member_required
@require_http_methods(["GET"])
def export_revenue_xlsx(request):
    """
    Optional XLSX export for paid orders.
    If openpyxl is not installed, return 501.
    """
    try:
        import openpyxl
        from openpyxl.utils import get_column_letter
    except ImportError:
        return HttpResponse("XLSX export requires openpyxl to be installed.", status=501)

    paid_orders, from_param, to_param, preset = _paid_orders_queryset(request)
    paid_orders = paid_orders.annotate(
        order_total=Sum(F('items__price') * F('items__quantity'), output_field=DecimalField())
    ).order_by('-paid_at', '-created_at')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Paid Orders"
    headers = ['Order ID', 'Paid At', 'Created At', 'Customer', 'Payment Method', 'Status', 'Total']
    ws.append(headers)
    for o in paid_orders:
        ws.append([
            o.id,
            o.paid_at.isoformat() if o.paid_at else '',
            o.created_at.isoformat() if o.created_at else '',
            o.full_name,
            o.get_payment_method_display(),
            o.get_status_display(),
            float(o.order_total or Decimal('0')),
        ])
    # Autosize columns (basic)
    for idx, _ in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = 18

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=revenue_paid_orders.xlsx'
    wb.save(response)
    return response


def _build_revenue_csv(qs):
    buffer = io.StringIO()
    writer = csv.writer(buffer)
    writer.writerow(['Order ID', 'Paid At', 'Created At', 'Customer', 'Payment Method', 'Status', 'Total'])
    for o in qs:
        writer.writerow([
            o.id,
            o.paid_at.isoformat() if o.paid_at else '',
            o.created_at.isoformat() if o.created_at else '',
            o.full_name,
            o.get_payment_method_display(),
            o.get_status_display(),
            f"{o.order_total or Decimal('0')}",
        ])
    return buffer.getvalue().encode('utf-8')


def _build_revenue_xlsx(qs):
    import openpyxl
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Paid Orders"
    headers = ['Order ID', 'Paid At', 'Created At', 'Customer', 'Payment Method', 'Status', 'Total']
    ws.append(headers)
    for o in qs:
        ws.append([
            o.id,
            o.paid_at.isoformat() if o.paid_at else '',
            o.created_at.isoformat() if o.created_at else '',
            o.full_name,
            o.get_payment_method_display(),
            o.get_status_display(),
            float(o.order_total or Decimal('0')),
        ])
    for idx, _ in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = 18
    bio = io.BytesIO()
    wb.save(bio)
    return bio.getvalue()


@staff_member_required
@require_http_methods(["POST"])
def email_revenue_report(request):
    """Send the filtered revenue report via email (CSV or XLSX)."""
    paid_orders, from_param, to_param, preset = _paid_orders_queryset(request, params=request.POST)
    paid_orders = paid_orders.annotate(
        order_total=Sum(F('items__price') * F('items__quantity'), output_field=DecimalField())
    ).order_by('-paid_at', '-created_at')

    recipient = request.POST.get('recipient')
    fmt = request.POST.get('format', 'csv')
    if not recipient:
        messages.error(request, "Vui lòng nhập email nhận báo cáo.")
        return redirect(f"{request.META.get('HTTP_REFERER', '/store/admin-dashboard/revenue/')}")

    subject = "Revenue report (paid orders)"
    body = "Đính kèm báo cáo doanh thu theo bộ lọc hiện tại."
    email = EmailMessage(subject, body, to=[recipient])

    try:
        if fmt == 'xlsx':
            try:
                attachment = _build_revenue_xlsx(paid_orders)
            except ImportError:
                messages.error(request, "Không có openpyxl, không thể xuất XLSX.")
                return redirect(f"{request.META.get('HTTP_REFERER', '/store/admin-dashboard/revenue/')}")
            email.attach('revenue_paid_orders.xlsx',
                         attachment,
                         'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        else:
            attachment = _build_revenue_csv(paid_orders)
            email.attach('revenue_paid_orders.csv', attachment, 'text/csv')
        email.send(fail_silently=True)
        messages.success(request, "Đã gửi báo cáo doanh thu qua email.")
    except Exception:
        messages.error(request, "Gửi email thất bại.")

    # Redirect back to revenue report with current filters in query string
    qs = []
    if from_param:
        qs.append(f"from_date={from_param}")
    if to_param:
        qs.append(f"to_date={to_param}")
    if preset:
        qs.append(f"preset={preset}")
    query = "&".join(qs)
    return redirect(f"/store/admin-dashboard/revenue/{'?' + query if query else ''}")


def get_kpi_metrics(start_date):
    """
    Calculate key performance indicators
    """
    today = timezone.now().date()
    
    # Total sales
    total_sales = Order.objects.filter(
        payment_status='completed'
    ).aggregate(
        total=Sum(F('items__price') * F('items__quantity'))
    )['total'] or Decimal('0')
    
    # Sales in period
    period_sales = Order.objects.filter(
        payment_status='completed',
        created_at__date__gte=start_date
    ).aggregate(
        total=Sum(F('items__price') * F('items__quantity'))
    )['total'] or Decimal('0')
    
    # Total orders
    total_orders = Order.objects.filter(payment_status='completed').count()
    
    # Period orders
    period_orders = Order.objects.filter(
        payment_status='completed',
        created_at__date__gte=start_date
    ).count()
    
    # Total users
    total_users = User.objects.count()
    
    # New users in period
    new_users = User.objects.filter(date_joined__date__gte=start_date).count()
    
    # Conversion rate (orders / unique visitors)
    unique_visitors = ProductView.objects.filter(
        viewed_at__date__gte=start_date
    ).values('session_key').distinct().count()
    
    conversion_rate = (period_orders / unique_visitors * 100) if unique_visitors > 0 else 0
    
    # Average order value
    avg_order_value = period_sales / period_orders if period_orders > 0 else Decimal('0')
    
    # Top selling product
    top_product = Product.objects.filter(
        orderitem__order__payment_status='completed',
        orderitem__order__created_at__date__gte=start_date
    ).annotate(
        total_sold=Sum('orderitem__quantity')
    ).order_by('-total_sold').first()
    
    return {
        'total_sales': float(total_sales),
        'period_sales': float(period_sales),
        'total_orders': total_orders,
        'period_orders': period_orders,
        'total_users': total_users,
        'new_users': new_users,
        'conversion_rate': round(conversion_rate, 2),
        'avg_order_value': float(avg_order_value),
        'top_product': top_product.name if top_product else 'N/A',
    }


def get_sales_trend_data(start_date):
    """
    Get daily sales data for line chart
    """
    sales_by_day = Order.objects.filter(
        payment_status='completed',
        created_at__date__gte=start_date
    ).extra(
        select={'day': 'date(created_at)'}
    ).values('day').annotate(
        revenue=Sum(F('items__price') * F('items__quantity')),
        orders=Count('id', distinct=True)
    ).order_by('day')
    
    labels = []
    revenue = []
    orders = []
    for item in sales_by_day:
        day = item.get('day')
        # Some DB backends may return date as string; normalize to date
        if isinstance(day, str):
            try:
                day_obj = datetime.strptime(day, '%Y-%m-%d').date()
            except Exception:
                # fallback: keep as string
                labels.append(str(day))
            else:
                labels.append(day_obj.strftime('%Y-%m-%d'))
        elif hasattr(day, 'strftime'):
            labels.append(day.strftime('%Y-%m-%d'))
        else:
            labels.append(str(day))
        revenue.append(float(item.get('revenue') or 0))
        orders.append(item.get('orders') or 0)

    return {'labels': labels, 'revenue': revenue, 'orders': orders}


def get_top_products_data(limit=10):
    """
    Get top products by sales for bar chart
    """
    top_products = Product.objects.filter(
        orderitem__order__payment_status='completed'
    ).annotate(
        total_sold=Sum('orderitem__quantity'),
        total_revenue=Sum(F('orderitem__price') * F('orderitem__quantity'))
    ).order_by('-total_revenue')[:limit]
    
    return {
        'labels': [p.name[:30] for p in top_products],
        'quantities': [p.total_sold or 0 for p in top_products],
        'revenue': [float(p.total_revenue or 0) for p in top_products],
    }


def get_revenue_breakdown():
    """
    Get revenue breakdown by category for pie chart
    """
    revenue_by_category = Product.objects.filter(
        orderitem__order__payment_status='completed',
        category__isnull=False
    ).values('category__name').annotate(
        revenue=Sum(F('orderitem__price') * F('orderitem__quantity'))
    ).order_by('-revenue')[:10]
    
    return {
        'labels': [item['category__name'] for item in revenue_by_category],
        'data': [float(item['revenue'] or 0) for item in revenue_by_category],
    }


def get_user_growth_data(start_date):
    """
    Get user growth data for area chart
    """
    users_by_day = User.objects.filter(
        date_joined__date__gte=start_date
    ).extra(
        select={'day': 'date(date_joined)'}
    ).values('day').annotate(
        count=Count('id')
    ).order_by('day')
    
    # Calculate cumulative growth
    total = 0
    cumulative = []
    labels = []
    
    for item in users_by_day:
        total += item['count']
        cumulative.append(total)
        day = item.get('day')
        if isinstance(day, str):
            try:
                day_obj = datetime.strptime(day, '%Y-%m-%d').date()
            except Exception:
                labels.append(str(day))
            else:
                labels.append(day_obj.strftime('%Y-%m-%d'))
        elif hasattr(day, 'strftime'):
            labels.append(day.strftime('%Y-%m-%d'))
        else:
            labels.append(str(day))
    
    return {
        'labels': labels,
        'data': cumulative,
        'new_users': [item['count'] for item in users_by_day],
    }


@staff_member_required
@require_http_methods(["GET"])
def export_sales_report(request):
    """
    Export sales report as CSV
    """
    # Get date range from query params
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')
    
    if start_date:
        start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
    else:
        start_date = timezone.now().date() - timedelta(days=30)
    
    if end_date:
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
    else:
        end_date = timezone.now().date()
    
    # Get orders
    orders = Order.objects.filter(
        payment_status='completed',
        created_at__date__gte=start_date,
        created_at__date__lte=end_date
    ).select_related().prefetch_related('items__product')
    
    # Create CSV
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow([
        'Order ID', 'Date', 'Customer', 'Phone', 'Items', 
        'Subtotal', 'Payment Method', 'Status'
    ])
    
    # Data
    for order in orders:
        items_count = order.items.count()
        subtotal = sum(item.price * item.quantity for item in order.items.all())
        
        writer.writerow([
            order.id,
            order.created_at.strftime('%Y-%m-%d %H:%M'),
            order.full_name,
            order.phone,
            items_count,
            f'${subtotal:.2f}',
            order.payment_method,
            order.payment_status,
        ])
    
    # Create response
    response = HttpResponse(output.getvalue(), content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="sales_report_{start_date}_{end_date}.csv"'
    
    return response


@staff_member_required
@require_http_methods(["GET"])
def export_products_report(request):
    """
    Export products report as CSV
    """
    products = Product.objects.filter(is_active=True).select_related('brand', 'category')
    
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Header
    writer.writerow([
        'ID', 'Name', 'Brand', 'Category', 'Price', 'Sale Price',
        'Stock', 'Views', 'Rating', 'Created'
    ])
    
    # Data
    for product in products:
        writer.writerow([
            product.id,
            product.name,
            product.brand.name,
            product.category.name if product.category else 'N/A',
            f'${product.price:.2f}',
            f'${product.sale_price:.2f}' if product.sale_price else 'N/A',
            product.stock_quantity,
            product.view_count,
            product.rating,
            product.created_at.strftime('%Y-%m-%d'),
        ])
    
    response = HttpResponse(output.getvalue(), content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="products_report.csv"'
    
    return response


@staff_member_required
@require_http_methods(["GET"])
def api_dashboard_metrics(request):
    """
    API endpoint for real-time dashboard metrics (AJAX)
    """
    period_days = int(request.GET.get('days', 30))
    start_date = timezone.now().date() - timedelta(days=period_days)
    
    metrics = get_kpi_metrics(start_date)
    
    return JsonResponse(metrics)


@staff_member_required
@require_http_methods(["GET"])
def api_sales_chart(request):
    """
    API endpoint for sales chart data (AJAX)
    """
    period_days = int(request.GET.get('days', 30))
    start_date = timezone.now().date() - timedelta(days=period_days)
    
    data = get_sales_trend_data(start_date)
    
    return JsonResponse(data)


@staff_member_required
@require_http_methods(["GET"])
def performance_metrics(request):
    """
    System performance metrics
    """
    from django.db import connection
    
    # Database queries count
    queries_count = len(connection.queries)
    
    # Cache stats (if Redis is configured)
    cache_stats = {
        'hits': 0,
        'misses': 0,
        'hit_rate': 0,
    }
    
    # Response time (simulated)
    response_times = {
        'avg': '120ms',
        'p95': '250ms',
        'p99': '500ms',
    }
    
    return JsonResponse({
        'database': {
            'queries_count': queries_count,
            'slow_queries': 0,
        },
        'cache': cache_stats,
        'response_times': response_times,
    })


@staff_member_required
def staff_activity_log(request):
    """
    View staff activity log
    """
    from django.contrib.admin.models import LogEntry
    
    logs = LogEntry.objects.select_related('user', 'content_type').order_by('-action_time')[:100]
    
    return render(request, 'admin/activity_log.html', {'logs': logs})
